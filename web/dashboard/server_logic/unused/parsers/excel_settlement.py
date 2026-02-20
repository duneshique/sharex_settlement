"""
Excel 정산 데이터 파서
==============================
Master 파일의 정산(실비) 시트에서 매출 데이터를,
Info 파일의 정산서(논리)/광고비 사용내역 시트에서 광고비 데이터를 파싱합니다.
"""

import re
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl

from .base import (
    CampaignCost,
    CourseSales,
    ParsedSettlementData,
    clean_numeric,
    load_course_mapping,
    get_course_company_id,
)


def parse_master_settlement_sheet(
    excel_path: str,
    sheet_name: str,
    month: str,
    base_path: str,
) -> List[CourseSales]:
    """
    Master 파일의 정산(실비) 시트에서 매출 데이터 추출

    시트 구조:
    - Row 1~18: 메타데이터 (회사정보, 작성일 등)
    - Row 19~20: 헤더 (정산월, 코스아이디, 강의명, 매출액, ...)
    - Row 21+: 데이터 행
    - 마지막: 합계 행

    Args:
        excel_path: Master Excel 파일 경로
        sheet_name: 시트명 (예: "24.10_정산(실비)")
        month: 정산월 (예: "2024-10")
        base_path: 프로젝트 루트

    Returns:
        List[CourseSales]
    """
    mapping = load_course_mapping(base_path)
    wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)

    if sheet_name not in wb.sheetnames:
        print(f"[Warning] 시트 '{sheet_name}'을 찾을 수 없습니다. 유사 시트 검색...")
        sheet_name = _find_similar_sheet(wb.sheetnames, month)
        if not sheet_name:
            wb.close()
            return []

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # 헤더 행 찾기 (동적 감지)
    header_row_idx, col_map = _detect_header(rows)
    if header_row_idx is None:
        print(f"[Warning] 헤더를 찾을 수 없습니다: {sheet_name}")
        return []

    print(f"  [{sheet_name}] 헤더 행: {header_row_idx + 1}, 컬럼: {col_map}")

    # 데이터 행 파싱
    sales = []
    for row_idx in range(header_row_idx + 1, len(rows)):
        row = rows[row_idx]
        if not row or len(row) <= max(col_map.values()):
            continue

        # 합계/소계 행 스킵
        row_text = " ".join(str(c or "") for c in row)
        if "합계" in row_text or "소계" in row_text:
            continue

        # 빈 행 스킵
        if all(c is None or str(c).strip() == "" for c in row):
            continue

        # 코스아이디 추출
        course_id_raw = row[col_map.get("course_id", 2)]
        if course_id_raw is None:
            continue

        course_id = str(course_id_raw).strip()
        # 숫자가 아닌 경우 스킵
        if not re.match(r'^\d{5,6}$', course_id.replace(".0", "")):
            continue
        course_id = course_id.replace(".0", "")

        # 매출액 추출
        revenue_raw = row[col_map.get("revenue", 5)]
        revenue = clean_numeric(revenue_raw)
        if revenue is None:
            revenue = 0.0

        # 강의명
        course_name_raw = row[col_map.get("course_name", 3)]
        course_name = str(course_name_raw or "").strip()

        # company_id 매핑
        company_id = get_course_company_id(course_id, mapping) or "unknown"

        sales.append(CourseSales(
            month=month,
            course_id=course_id,
            course_name=course_name,
            company_id=company_id,
            revenue=revenue,
        ))

    return sales


def parse_info_campaign_data(
    excel_path: str,
    period_months: List[str],
    base_path: str,
) -> List[CampaignCost]:
    """
    Info 파일의 정산서(논리) 시트에서 광고비 데이터 추출

    이 시트에는 캠페인별 분류가 이미 되어 있음:
    - target: SHARE X / PLUS X / BKID / BLSN / SANDOLL
    - channel: Meta / Google / Naver

    Args:
        excel_path: Info Excel 파일 경로
        period_months: 대상 월 리스트 (예: ["2024-10", "2024-11", "2024-12"])
        base_path: 프로젝트 루트

    Returns:
        List[CampaignCost]
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    costs = []

    # 정산서(논리) 시트 탐색
    logic_sheet = None
    for name in wb.sheetnames:
        if "정산서" in name and "논리" in name:
            logic_sheet = name
            break

    if logic_sheet:
        ws = wb[logic_sheet]
        rows = list(ws.iter_rows(values_only=True))
        costs.extend(_parse_logic_sheet(rows, period_months))

    # 광고비 사용내역 시트도 탐색
    ad_sheet = None
    for name in wb.sheetnames:
        if "광고비" in name and "사용내역" in name:
            ad_sheet = name
            break

    if ad_sheet:
        ws = wb[ad_sheet]
        rows = list(ws.iter_rows(values_only=True))
        ad_costs = _parse_ad_usage_sheet(rows, period_months)
        if ad_costs and not costs:
            costs = ad_costs

    wb.close()
    return costs


def parse_info_raw_data(
    excel_path: str,
    period_months: List[str],
    base_path: str,
) -> List[CourseSales]:
    """
    Info 파일의 정산데이터(Raw) 시트에서 매출 데이터 추출

    이 시트에는 2023.01~2025.06 전체 원본 데이터가 있음.
    코스별/월별 매출액을 추출.

    Args:
        excel_path: Info Excel 파일 경로
        period_months: 대상 월 리스트
        base_path: 프로젝트 루트

    Returns:
        List[CourseSales]
    """
    mapping = load_course_mapping(base_path)
    wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)

    raw_sheet = None
    for name in wb.sheetnames:
        if "정산데이터" in name and "Raw" in name:
            raw_sheet = name
            break

    if not raw_sheet:
        wb.close()
        return []

    ws = wb[raw_sheet]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # 헤더 찾기
    header_idx = None
    col_map = {}
    for i, row in enumerate(rows):
        row_text = " ".join(str(c or "") for c in row).lower()
        if "코스" in row_text and ("매출" in row_text or "정산" in row_text):
            header_idx = i
            for j, cell in enumerate(row):
                cell_str = str(cell or "").strip()
                if "코스" in cell_str and ("아이디" in cell_str or "id" in cell_str.lower()):
                    col_map["course_id"] = j
                elif "강의명" in cell_str or "코스명" in cell_str:
                    col_map["course_name"] = j
                elif "매출" in cell_str:
                    col_map["revenue"] = j
                elif "정산월" in cell_str or "월" in cell_str:
                    col_map["month"] = j
                elif "회사" in cell_str or "기업" in cell_str:
                    col_map["company"] = j
            break

    if header_idx is None:
        return []

    sales = []
    for row_idx in range(header_idx + 1, len(rows)):
        row = rows[row_idx]
        if not row:
            continue

        # 월 추출
        month_raw = row[col_map.get("month", 0)]
        if month_raw is None:
            continue

        month_str = str(month_raw).strip()
        # datetime 객체 처리
        if hasattr(month_raw, 'strftime'):
            month_str = month_raw.strftime("%Y-%m")
        else:
            parsed = _parse_month_value(month_str)
            if parsed:
                month_str = parsed
            else:
                continue

        if month_str not in period_months:
            continue

        # 코스ID
        course_id_raw = row[col_map.get("course_id", 1)]
        if course_id_raw is None:
            continue
        course_id = str(course_id_raw).strip().replace(".0", "")
        if not re.match(r'^\d{5,6}$', course_id):
            continue

        # 매출액
        revenue_raw = row[col_map.get("revenue", 3)]
        revenue = clean_numeric(revenue_raw)
        if revenue is None:
            revenue = 0.0

        # 강의명
        course_name = str(row[col_map.get("course_name", 2)] or "").strip()

        # company_id
        company_id = get_course_company_id(course_id, mapping) or "unknown"

        sales.append(CourseSales(
            month=month_str,
            course_id=course_id,
            course_name=course_name,
            company_id=company_id,
            revenue=revenue,
        ))

    return sales


# ──────────────────────────────────────────────
# 내부 헬퍼 함수
# ──────────────────────────────────────────────

def _detect_header(rows: list) -> Tuple[Optional[int], Dict[str, int]]:
    """
    동적 헤더 감지: row 10~55 범위에서 "코스아이디"/"강의명"/"매출액" 키워드 탐색
    """
    for i in range(min(10, len(rows)), min(55, len(rows))):
        row = rows[i]
        if not row:
            continue

        row_text = " ".join(str(c or "") for c in row)

        if "코스아이디" in row_text or "코스 아이디" in row_text:
            col_map = {}
            for j, cell in enumerate(row):
                cell_str = str(cell or "").strip()
                if "코스" in cell_str and ("아이디" in cell_str or "ID" in cell_str):
                    col_map["course_id"] = j
                elif "강의명" in cell_str:
                    col_map["course_name"] = j
                elif "매출액" in cell_str or "매출" == cell_str:
                    col_map["revenue"] = j
                elif "직접" in cell_str and "광고" in cell_str:
                    col_map["direct_ad"] = j
                elif "간접" in cell_str and "광고" in cell_str:
                    col_map["indirect_ad"] = j
                elif "마케팅" in cell_str:
                    col_map["marketing"] = j
                elif "공헌이익" in cell_str:
                    col_map["contribution"] = j
                elif "강사료" in cell_str or "수익쉐어" in cell_str:
                    col_map["rs_fee"] = j
                elif "정산월" in cell_str:
                    col_map["month"] = j

            return i, col_map

    return None, {}


def _find_similar_sheet(sheet_names: list, month: str) -> Optional[str]:
    """월 기준으로 유사 시트명 찾기"""
    # "2024-10" → "24.10"
    year_short = month[2:4]
    month_num = month[5:7]
    pattern = f"{year_short}.{month_num}"

    for name in sheet_names:
        if pattern in name and "정산" in name:
            return name

    return None


def _parse_logic_sheet(rows: list, period_months: List[str]) -> List[CampaignCost]:
    """정산서(논리) 시트 파싱"""
    costs = []

    # 헤더 찾기
    header_idx = None
    col_map = {}
    for i, row in enumerate(rows):
        if not row:
            continue
        row_text = " ".join(str(c or "") for c in row)
        if ("채널" in row_text or "매체" in row_text) and ("광고비" in row_text or "비용" in row_text):
            header_idx = i
            for j, cell in enumerate(row):
                cell_str = str(cell or "").strip()
                if "채널" in cell_str or "매체" in cell_str:
                    col_map["channel"] = j
                elif "대상" in cell_str or "타겟" in cell_str or "target" in cell_str.lower():
                    col_map["target"] = j
                elif "캠페인" in cell_str:
                    col_map["campaign"] = j
                elif "광고비" in cell_str or "비용" in cell_str or "금액" in cell_str:
                    col_map["cost"] = j
                elif "월" in cell_str or "기간" in cell_str:
                    col_map["month"] = j
            break

    if header_idx is None:
        return costs

    for row_idx in range(header_idx + 1, len(rows)):
        row = rows[row_idx]
        if not row:
            continue

        # 빈 행 스킵
        if all(c is None or str(c).strip() == "" for c in row):
            continue

        channel = str(row[col_map.get("channel", 0)] or "").strip()
        target = str(row[col_map.get("target", 1)] or "").strip()
        campaign = str(row[col_map.get("campaign", 2)] or "").strip()
        cost_raw = row[col_map.get("cost", 3)]
        month_raw = row[col_map.get("month", 4)] if "month" in col_map else None

        cost = clean_numeric(cost_raw)
        if cost is None or cost == 0:
            continue

        # 월 처리
        if month_raw:
            month_str = _parse_month_value(str(month_raw))
            if month_str and month_str not in period_months:
                continue

        costs.append(CampaignCost(
            month=month_str if month_raw else period_months[0],
            channel=channel,
            target=target.upper() if target else "SHARE X",
            campaign_name=campaign,
            cost_krw=cost,
        ))

    return costs


def _parse_ad_usage_sheet(rows: list, period_months: List[str]) -> List[CampaignCost]:
    """광고비 사용내역 시트 파싱"""
    # 이 시트는 Info 파일에 따라 구조가 다를 수 있으므로 범용적으로 파싱
    costs = []

    for i, row in enumerate(rows):
        if not row:
            continue

        row_text = " ".join(str(c or "") for c in row)

        # 월별 합계 행 찾기 (예: "2024.10", "10월")
        month_match = re.search(r'20(\d{2})[.\-](\d{1,2})', row_text)
        if not month_match:
            continue

        month_str = f"20{month_match.group(1)}-{int(month_match.group(2)):02d}"
        if month_str not in period_months:
            continue

        # 숫자 추출
        nums = []
        for cell in row:
            val = clean_numeric(cell)
            if val is not None and val > 0:
                nums.append(val)

        if nums:
            costs.append(CampaignCost(
                month=month_str,
                channel="Total",
                target="SHARE X",
                campaign_name=f"광고비 합계 {month_str}",
                cost_krw=nums[0],
            ))

    return costs


def _parse_month_value(text: str) -> Optional[str]:
    """다양한 월 형식을 YYYY-MM으로 변환"""
    # "2024-10" 이미 정규 형식
    if re.match(r'^\d{4}-\d{2}$', text):
        return text

    # "2024.10" 또는 "2024년 10월"
    m = re.search(r'(\d{4})[.\-년]\s*(\d{1,2})', text)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}"

    # datetime 문자열 "2024-10-31T00:00:00"
    m = re.match(r'(\d{4})-(\d{2})', text)
    if m:
        return f"{m.group(1)}-{m.group(2)}"

    return None
