"""
Share X 정산 마스터 엑셀 - 데이터 구조 분석기
==============================================
업무용 엑셀 파일의 시트별 구조를 자동으로 분석하여
데이터 전처리 및 자동화 워크플로우 설계를 위한 기초 정보를 제공합니다.

입력: 엑셀 파일 경로
출력: 각 시트의 구조 정보 (시트 유형, 헤더 위치, 컬럼명, 데이터 타입, 유효 행수)
"""

import openpyxl
import json
import re
import sys
from datetime import datetime
from collections import Counter, OrderedDict


# ── 시트 유형 분류 규칙 ──
SHEET_TYPE_RULES = {
    "정산(실비)": {
        "pattern": r"정산\(실비\)|정산\(기존방식\)|정산\(실제광고비\)|정산\(실비최대",
        "description": "월별 수익쉐어 정산내역서",
        "expected_header_row_range": (19, 25),
    },
    "정산_통합": {
        "pattern": r"정산.*통합|정산\(통합\)",
        "description": "분기별 정산 통합본",
        "expected_header_row_range": (19, 25),
    },
    "광고": {
        "pattern": r"광고",
        "description": "광고비 사용내역 (Adriel 기반)",
        "expected_header_row_range": (9, 12),
    },
    "컨퍼런스_제작비": {
        "pattern": r"컨퍼런스 제작비",
        "description": "컨퍼런스 지출 상세 내역",
        "expected_header_row_range": (3, 5),
    },
    "B2B_판매리스트": {
        "pattern": r"B2B",
        "description": "B2B 판매가 및 할인율 마스터",
        "expected_header_row_range": (1, 2),
    },
    "재계산": {
        "pattern": r"재계산",
        "description": "정산 재계산 검증용",
        "expected_header_row_range": (1, 1),
    },
    "허스키": {
        "pattern": r"허스키",
        "description": "허스키폭스 별도 정산",
        "expected_header_row_range": (19, 25),
    },
    "기타": {
        "pattern": r"시트\d+",
        "description": "임시/작업 시트",
        "expected_header_row_range": (1, 5),
    },
}


def classify_sheet(sheet_name):
    """시트명으로 유형 분류"""
    for stype, rule in SHEET_TYPE_RULES.items():
        if re.search(rule["pattern"], sheet_name):
            return stype, rule["description"]
    return "미분류", "분류되지 않은 시트"


def detect_header_row(rows_data, sheet_type):
    """헤더 행 자동 감지: 텍스트 셀이 가장 많은 행을 헤더로 판단"""
    best_row = None
    best_score = 0
    for i, row in enumerate(rows_data):
        text_count = sum(1 for v in row if isinstance(v, str) and len(str(v).strip()) > 1)
        if text_count > best_score:
            best_score = text_count
            best_row = i
    return best_row


def infer_column_type(values):
    """셀 값 리스트에서 실제 데이터 타입 추론"""
    type_counts = Counter()
    non_null = 0
    for v in values:
        if v is None or (isinstance(v, str) and v.strip() in ("", "-", "N/A")):
            continue
        non_null += 1
        if isinstance(v, datetime):
            type_counts["datetime"] += 1
        elif isinstance(v, bool):
            type_counts["bool"] += 1
        elif isinstance(v, (int, float)):
            type_counts["numeric"] += 1
        elif isinstance(v, str):
            # 숫자 문자열 체크
            cleaned = v.replace(",", "").replace("₩", "").replace("%", "").strip()
            try:
                float(cleaned)
                type_counts["numeric_string"] += 1
            except ValueError:
                type_counts["text"] += 1
        else:
            type_counts["other"] += 1

    if non_null == 0:
        return "empty", 0

    dominant = type_counts.most_common(1)[0][0] if type_counts else "empty"
    return dominant, non_null


def parse_period_from_name(sheet_name):
    """시트명에서 연월 정보 추출"""
    # 패턴: 26.1, 25.12, 24.01, 23.09 등
    m = re.search(r"(\d{2})\.(\d{1,2})", sheet_name)
    if m:
        year = int(m.group(1)) + 2000
        month = int(m.group(2))
        return f"{year}-{month:02d}"
    # 분기 패턴: 25.1Q, 24.4Q
    m = re.search(r"(\d{2})\.(\d)Q", sheet_name)
    if m:
        year = int(m.group(1)) + 2000
        quarter = int(m.group(2))
        return f"{year}-Q{quarter}"
    return None


def analyze_sheet(ws, sheet_name, max_scan_rows=50):
    """단일 시트 구조 분석"""
    sheet_type, description = classify_sheet(sheet_name)
    period = parse_period_from_name(sheet_name)

    # 전체 행 수 및 데이터 수집
    all_rows = []
    total_rows = 0
    for row in ws.iter_rows(values_only=True):
        total_rows += 1
        if len(all_rows) < max_scan_rows:
            all_rows.append(list(row))

    # 헤더 행 감지
    header_row_idx = detect_header_row(all_rows, sheet_type)
    if header_row_idx is None:
        return {
            "sheet_name": sheet_name,
            "sheet_type": sheet_type,
            "description": description,
            "period": period,
            "total_rows": total_rows,
            "status": "EMPTY_OR_UNREADABLE",
            "columns": [],
        }

    # 헤더 추출
    header_row = all_rows[header_row_idx]
    columns_info = []
    data_rows = all_rows[header_row_idx + 1:]

    for col_idx, col_name in enumerate(header_row):
        if col_name is None:
            continue
        col_name_str = str(col_name).strip().replace("\n", " ")
        if not col_name_str or col_name_str == "-":
            continue

        # 해당 컬럼의 데이터 값 수집
        col_values = []
        for drow in data_rows:
            if col_idx < len(drow):
                col_values.append(drow[col_idx])

        dtype, non_null_count = infer_column_type(col_values)
        data_row_count = len(data_rows)

        # 샘플 값 (처음 3개 non-null)
        samples = []
        for v in col_values:
            if v is not None and str(v).strip() not in ("", "-"):
                samples.append(str(v)[:50])
                if len(samples) >= 3:
                    break

        columns_info.append({
            "col_index": col_idx + 1,
            "col_name": col_name_str,
            "data_type": dtype,
            "non_null_count": non_null_count,
            "null_ratio": round(1 - non_null_count / max(data_row_count, 1), 2),
            "samples": samples,
        })

    # 데이터 시작 행 (엑셀 기준, 1-indexed)
    data_start_row = header_row_idx + 2  # 헤더 다음 행

    # 유효 데이터 행 수 (빈 행 제외)
    valid_data_rows = 0
    for drow in data_rows:
        if any(v is not None and str(v).strip() not in ("", "-") for v in drow):
            valid_data_rows += 1

    return {
        "sheet_name": sheet_name,
        "sheet_type": sheet_type,
        "description": description,
        "period": period,
        "total_rows": total_rows,
        "header_row": header_row_idx + 1,
        "data_start_row": data_start_row,
        "valid_data_rows": valid_data_rows,
        "total_columns": len(columns_info),
        "columns": columns_info,
    }


def analyze_workbook(file_path):
    """워크북 전체 구조 분석"""
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

    results = {
        "file_path": file_path,
        "analysis_date": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "total_sheets": len(wb.sheetnames),
        "sheet_type_summary": {},
        "sheets": [],
    }

    type_counter = Counter()
    for sname in wb.sheetnames:
        ws = wb[sname]
        sheet_info = analyze_sheet(ws, sname)
        results["sheets"].append(sheet_info)
        type_counter[sheet_info["sheet_type"]] += 1

    wb.close()

    results["sheet_type_summary"] = dict(type_counter.most_common())
    return results


def print_report(results):
    """분석 결과 리포트 출력"""
    print("=" * 80)
    print(f"  📊 엑셀 데이터 구조 분석 리포트")
    print(f"  파일: {results['file_path']}")
    print(f"  분석일시: {results['analysis_date']}")
    print(f"  총 시트 수: {results['total_sheets']}")
    print("=" * 80)

    # 시트 유형 요약
    print("\n┌─ 시트 유형별 분포 ─────────────────────────────────┐")
    for stype, count in results["sheet_type_summary"].items():
        desc = ""
        for rule in SHEET_TYPE_RULES.values():
            if stype in SHEET_TYPE_RULES:
                desc = SHEET_TYPE_RULES[stype]["description"]
                break
        print(f"  │ {stype:<20} : {count:>3}개  ({desc})")
    print("└──────────────────────────────────────────────────────┘")

    # 시트별 상세
    for sheet in results["sheets"]:
        print(f"\n{'─' * 80}")
        period_str = f" [{sheet['period']}]" if sheet.get("period") else ""
        print(f"📋 {sheet['sheet_name']}{period_str}")
        print(f"   유형: {sheet['sheet_type']} | {sheet['description']}")

        if sheet.get("status") == "EMPTY_OR_UNREADABLE":
            print(f"   ⚠️  상태: 비어있거나 읽을 수 없는 시트 (총 {sheet['total_rows']}행)")
            continue

        print(f"   총 행: {sheet['total_rows']} | 헤더 행: {sheet['header_row']} | "
              f"데이터 시작: {sheet['data_start_row']} | 유효 데이터: {sheet['valid_data_rows']}행")
        print(f"   컬럼 수: {sheet['total_columns']}")

        if sheet["columns"]:
            print(f"   {'─' * 74}")
            print(f"   {'#':<4} {'컬럼명':<35} {'타입':<16} {'비공란':<8} {'NULL%':<6}")
            print(f"   {'─' * 74}")
            for col in sheet["columns"]:
                print(f"   {col['col_index']:<4} {col['col_name'][:34]:<35} "
                      f"{col['data_type']:<16} {col['non_null_count']:<8} "
                      f"{col['null_ratio']:.0%}")
                if col.get("samples"):
                    sample_str = " | ".join(col["samples"][:2])
                    print(f"        └ 예시: {sample_str[:65]}")

    # 전처리 권고사항
    print(f"\n{'=' * 80}")
    print("  🔧 데이터 전처리 권고사항")
    print("=" * 80)

    issues = []
    # 헤더 위치 불일치 검출
    header_positions = {}
    for sheet in results["sheets"]:
        st = sheet["sheet_type"]
        hr = sheet.get("header_row")
        if hr:
            header_positions.setdefault(st, set()).add(hr)

    for st, positions in header_positions.items():
        if len(positions) > 1:
            issues.append(
                f"[구조 불일치] '{st}' 유형 시트들의 헤더 행 위치가 다릅니다: {sorted(positions)}행\n"
                f"   → 시트별 헤더 행 위치를 동적으로 감지하는 로직 필요"
            )

    # 컬럼명 불일치 검출
    type_columns = {}
    for sheet in results["sheets"]:
        st = sheet["sheet_type"]
        cols = tuple(c["col_name"] for c in sheet.get("columns", []))
        if cols:
            type_columns.setdefault(st, []).append((sheet["sheet_name"], cols))

    for st, sheet_cols in type_columns.items():
        if len(sheet_cols) > 1:
            unique_schemas = set(cols for _, cols in sheet_cols)
            if len(unique_schemas) > 1:
                issues.append(
                    f"[스키마 변화] '{st}' 유형의 컬럼 구성이 시트마다 다릅니다 ({len(unique_schemas)}가지 변형)\n"
                    f"   → 컬럼 매핑 테이블을 작성하여 정규화 필요"
                )

    # 날짜 형식 혼재
    issues.append(
        "[데이터 타입] 날짜 컬럼이 datetime과 text가 혼재합니다\n"
        "   → pd.to_datetime() 파싱 시 errors='coerce' 옵션 사용 권장"
    )
    issues.append(
        "[문서 헤더] 정산 시트 상단 1~18행은 회사정보/메타데이터로 데이터 로딩 시 skiprows 처리 필요\n"
        "   → 시트 유형별 skiprows 매핑: 정산=~19행, 광고=~10행, B2B=0행"
    )
    issues.append(
        "[병합 셀] 업무용 엑셀 특성상 셀 병합이 다수 존재하며 pandas 로딩 시 NaN으로 처리됨\n"
        "   → forward fill (ffill) 처리 필요"
    )
    issues.append(
        "[금액 데이터] 일부 금액 셀에 '-' 문자열이 0원 대신 사용됨\n"
        "   → '-' → 0 변환 후 numeric 캐스팅 필요"
    )

    for i, issue in enumerate(issues, 1):
        print(f"\n  {i}. {issue}")

    print(f"\n{'=' * 80}")


def save_json_report(results, output_path):
    """분석 결과를 JSON으로 저장"""
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2, default=str)
    print(f"\n✅ JSON 리포트 저장: {output_path}")


if __name__ == "__main__":
    file_path = sys.argv[1] if len(sys.argv) > 1 else "/mnt/project/Share_X_Settlement_Master_20232026.xlsx"
    json_output = sys.argv[2] if len(sys.argv) > 2 else "/home/claude/structure_analysis.json"

    results = analyze_workbook(file_path)
    print_report(results)
    save_json_report(results, json_output)
