#!/usr/bin/env python3
"""
Campaign Classification Rules Extractor
========================================

This script extracts campaign classification rules from the ShareX Settlement Excel file
and generates a JSON configuration file.

Source: Share X Settlement_Info.xlsx
Target Sheet: 정산서(논리)
Output: config/campaign_rules.json

Author: Claude Code
Date: 2026-02-09
"""

import openpyxl
from openpyxl import load_workbook
import json
from datetime import datetime
from pathlib import Path


def extract_campaign_rules(excel_path, output_dir=None):
    """
    Extract campaign classification rules from Excel file.

    Args:
        excel_path (str or Path): Path to the Excel file
        output_dir (str or Path, optional): Output directory. Defaults to same as excel_path/config

    Returns:
        dict: Campaign rules configuration
    """

    # Determine paths
    excel_path = Path(excel_path)
    if output_dir is None:
        output_dir = excel_path.parent / "config"
    else:
        output_dir = Path(output_dir)

    output_file = output_dir / "campaign_rules.json"

    # Create output directory if it doesn't exist
    output_dir.mkdir(parents=True, exist_ok=True)

    # Load the workbook
    wb = load_workbook(excel_path, data_only=True)

    # Extract from 정산서(논리) sheet
    sheet = wb["정산서(논리)"]

    # Get course counts from Row 1 and 2
    plusx_courses = 24.0
    union_courses = 15.0

    for col_idx in range(11, 13):
        label_k = sheet.cell(row=1, column=11).value
        label_l = sheet.cell(row=2, column=11).value
        if label_k == "플엑 강의 수":
            plusx_courses = float(sheet.cell(row=1, column=12).value or 24)
        if label_l == "유니온 강의 수":
            union_courses = float(sheet.cell(row=2, column=12).value or 15)

    total_courses = plusx_courses + union_courses

    # Extract data from Row 4 onwards (Row 3 is header)
    unique_ad_targets = set()
    unique_campaign_names = set()
    unique_channels = set()
    exchange_rates = {}

    for row_idx in range(4, sheet.max_row + 1):
        # Column A: 분기
        quarter = sheet.cell(row=row_idx, column=1).value
        if not quarter:
            break

        # Column B: 월 (Month) - extract date
        month_cell = sheet.cell(row=row_idx, column=2).value
        year_month = None
        if month_cell:
            # Extract date
            if isinstance(month_cell, str):
                # Parse date string like "2024-10-01 00:00:00"
                date_str = month_cell.split()[0]  # Get "2024-10-01"
                year_month = date_str[:7]  # Get "2024-10"
            else:
                # It's a datetime object
                year_month = month_cell.strftime("%Y-%m")

        # Column C: 채널 (Channel)
        channel = sheet.cell(row=row_idx, column=3).value
        if channel:
            unique_channels.add(channel)

        # Column D: 광고 대상 (Ad Target)
        ad_target = sheet.cell(row=row_idx, column=4).value
        if ad_target:
            unique_ad_targets.add(ad_target)

        # Column E: 캠페인명 (Campaign Name)
        campaign_name = sheet.cell(row=row_idx, column=5).value
        if campaign_name:
            unique_campaign_names.add(campaign_name)

        # Column H: 적용환율 (Exchange Rate)
        exchange_rate = sheet.cell(row=row_idx, column=8).value
        if exchange_rate and year_month:
            if year_month not in exchange_rates:
                exchange_rates[year_month] = exchange_rate

    # Build the target mapping based on rules
    target_mapping = {
        "SHARE X": {
            "type": "indirect",
            "description": "통합 광고 - 전체 강의 수로 균등 안분",
            "apportionment": "equal_by_course_count",
            "note": "모든 회사의 강의를 합산한 총 강의 수로 나누어 배분"
        },
        "PLUS X": {
            "type": "direct",
            "company_id": "plusx",
            "description": "플러스엑스 직접광고비",
            "course_count": plusx_courses
        },
        "BKID": {
            "type": "direct",
            "company_id": "bkid",
            "description": "BKID 직접광고비"
        },
        "BLSN": {
            "type": "direct",
            "company_id": "blsn",
            "description": "블센 직접광고비"
        },
        "SANDOLL": {
            "type": "direct",
            "company_id": "sandoll",
            "description": "산돌 직접광고비"
        }
    }

    # Campaign patterns
    campaign_patterns = [
        {
            "pattern": "쉐어엑스|Share X|SHARE X|SA_",
            "default_target": "SHARE X",
            "category": "통합광고"
        },
        {
            "pattern": "플러스|PLUS|IP_",
            "default_target": "PLUS X",
            "category": "플러스엑스"
        },
        {
            "pattern": "BKID|비케이아이디|DA_",
            "default_target": "BKID",
            "category": "BKID"
        },
        {
            "pattern": "산돌|SANDOLL",
            "default_target": "SANDOLL",
            "category": "산돌"
        },
        {
            "pattern": "블센|BLSN",
            "default_target": "BLSN",
            "category": "블센"
        }
    ]

    # Build the comprehensive JSON object
    campaign_rules = {
        "classification_rules": {
            "target_mapping": target_mapping,
            "campaign_patterns": campaign_patterns
        },
        "channels": sorted(list(unique_channels)),
        "apportionment_formula": {
            "description": "광고비 = 총 광고비 A ÷ 전체 강의 수 B × 해당 기업 강의 수 C",
            "formula": "company_ad_cost = total_ad_cost / total_courses * company_courses",
            "note": "간접광고(SHARE X)는 이 공식으로 각 기업에 배분. 직접광고는 100% 해당 기업 귀속"
        },
        "course_count": {
            "plusx_courses": plusx_courses,
            "union_courses": union_courses,
            "total_courses": total_courses,
            "note": "시기별 변동 가능 (38~40개)"
        },
        "exchange_rates": {k: v for k, v in sorted(exchange_rates.items())},
        "unique_ad_targets": sorted(list(unique_ad_targets)),
        "unique_campaign_names": sorted(list(unique_campaign_names)),
        "metadata": {
            "source": "Share X Settlement_Info.xlsx > 정산서(논리), 광고비 사용내역",
            "extracted_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "total_unique_targets": len(unique_ad_targets),
            "total_unique_campaigns": len(unique_campaign_names),
            "total_unique_channels": len(unique_channels)
        }
    }

    # Write to JSON file
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(campaign_rules, f, ensure_ascii=False, indent=2)

    return campaign_rules, output_file


def main():
    """Main entry point."""
    file_path = Path(__file__).parent / "Share X Settlement_Info.xlsx"

    if not file_path.exists():
        # Try alternative path
        file_path = Path("/sessions/blissful-festive-dijkstra/mnt/ShareX_Settlement/Share X Settlement_Info.xlsx")

    if not file_path.exists():
        print(f"Error: Excel file not found at {file_path}")
        return

    print("Extracting campaign classification rules...")
    campaign_rules, output_file = extract_campaign_rules(file_path)

    print(f"✓ Campaign rules extracted successfully!")
    print(f"✓ Output file: {output_file}")
    print(f"\nExtracted Summary:")
    print(f"  - Ad Targets: {campaign_rules['unique_ad_targets']}")
    print(f"  - Channels: {campaign_rules['channels']}")
    print(f"  - Exchange Rates: {len(campaign_rules['exchange_rates'])} periods")
    print(f"  - Course Counts: PLUS X={campaign_rules['course_count']['plusx_courses']}, "
          f"Union={campaign_rules['course_count']['union_courses']}, "
          f"Total={campaign_rules['course_count']['total_courses']}")
    print(f"  - Unique Campaigns: {len(campaign_rules['unique_campaign_names'])}")


if __name__ == "__main__":
    main()
